from __future__ import annotations

import base64
import io
import json
import tempfile
from typing import List, Tuple
from pathlib import Path

import anthropic
import fitz  # PyMuPDF
from PIL import Image

from config import ANTHROPIC_API_KEY, CLAUDE_MODEL, PDF_DPI, MAX_IMAGE_WIDTH, MAX_IMAGE_HEIGHT
from models.guideline import ParsedGuideline

GUIDELINE_PARSE_PROMPT = """You are an expert at analyzing creator content guidelines/briefs for video production.

Analyze the provided guideline document images and extract ALL information into the following JSON structure.
Be thorough — capture every rule, scene direction, mandatory element, and suggestion.

Return ONLY valid JSON (no markdown, no explanation):

{
  "title": "Guideline title",
  "product_name": "Product name(s)",
  "concept": "Overall concept/theme of the content",
  "content_objective": "Campaign goal/objective",
  "video_duration": "Recommended video duration (e.g. '15-30 seconds')",
  "key_message": "Core message to convey",
  "rules": [
    {
      "category": "do|dont|brand_rule|mandatory",
      "description": "Rule description",
      "severity": "strict|recommended|optional"
    }
  ],
  "scenes": [
    {
      "scene_number": 1,
      "time_range": "0-3s or null",
      "description": "What should happen in this scene",
      "visual_direction": "Visual/filming direction",
      "script_suggestion": "Suggested script/voice-over lines",
      "text_overlay": "On-screen text to include or null"
    }
  ],
  "mandatory_elements": ["hashtags", "mentions", "specific visuals", etc.],
  "recommended_flow": "Summary of the overall recommended flow (e.g. Intro → Body → Outro)"
}

Important instructions:
- For "rules": Extract ALL do's, don'ts, brand rules, and mandatory requirements.
  - "brand_rule" with "strict" severity = language/tone restrictions (e.g. forbidden words)
  - "mandatory" = must-include elements
  - "do" = recommended practices
  - "dont" = things to avoid
- For "scenes": Extract every scene/cut direction. If there are timestamp ranges, include them.
  - Include visual directions (close-up, B&A, split screen, etc.)
  - Include suggested script lines and text overlays
- For "mandatory_elements": List specific things that MUST appear (hashtags, @mentions, specific shots, etc.)
- Capture EVERYTHING — missing a rule or scene could cause compliance issues.
- If information is not available for a field, use empty string "" or null as appropriate.
"""


def resize_image_bytes(image_bytes: bytes, max_w: int = MAX_IMAGE_WIDTH, max_h: int = MAX_IMAGE_HEIGHT) -> bytes:
    """Resize image to fit within max dimensions while maintaining aspect ratio."""
    img = Image.open(io.BytesIO(image_bytes))
    if img.mode in ("RGBA", "P", "LA"):
        img = img.convert("RGB")
    img.thumbnail((max_w, max_h), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return buf.getvalue()


def pdf_to_images(pdf_bytes: bytes, dpi: int = PDF_DPI) -> list[bytes]:
    """Convert PDF pages to JPEG images."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    zoom = dpi / 72
    matrix = fitz.Matrix(zoom, zoom)
    for page in doc:
        pix = page.get_pixmap(matrix=matrix)
        img_bytes = pix.tobytes("jpeg")
        img_bytes = resize_image_bytes(img_bytes)
        images.append(img_bytes)
    doc.close()
    return images


def excel_to_images(excel_bytes: bytes) -> tuple[list[bytes], str]:
    """Extract text and embedded images from Excel file.
    Returns (images, text_content)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    text_parts = []
    images = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"\n=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip(" |"):
                text_parts.append(row_text)

        # Extract embedded images
        for img in ws._images:
            img_bytes = img._data()
            if img_bytes:
                img_bytes = resize_image_bytes(img_bytes)
                images.append(img_bytes)

    return images, "\n".join(text_parts)


def image_to_bytes(uploaded_file) -> bytes:
    """Convert uploaded image file to resized JPEG bytes."""
    raw = uploaded_file.read() if hasattr(uploaded_file, "read") else uploaded_file
    return resize_image_bytes(raw)


def parse_guideline(
    files: list[tuple[str, bytes]],
) -> ParsedGuideline:
    """Parse guideline documents into structured data using Claude multimodal.

    Args:
        files: List of (filename, file_bytes) tuples.
               Supported: .pdf, .xlsx, .xls, .png, .jpg, .jpeg

    Returns:
        ParsedGuideline with extracted information.
    """
    all_images: list[bytes] = []
    extra_text = ""

    for filename, file_bytes in files:
        ext = Path(filename).suffix.lower()
        if ext == ".pdf":
            all_images.extend(pdf_to_images(file_bytes))
        elif ext in (".xlsx", ".xls"):
            imgs, text = excel_to_images(file_bytes)
            all_images.extend(imgs)
            extra_text += text + "\n"
        elif ext == ".csv":
            extra_text += file_bytes.decode("utf-8", errors="replace") + "\n"
        elif ext in (".png", ".jpg", ".jpeg", ".webp"):
            all_images.append(resize_image_bytes(file_bytes))
        else:
            raise ValueError(f"Unsupported file format: {ext}")

    if not all_images and not extra_text:
        raise ValueError("No content extracted from uploaded files.")

    # Build Claude API message content
    content = []

    for i, img_bytes in enumerate(all_images):
        b64 = base64.b64encode(img_bytes).decode("utf-8")
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/jpeg",
                "data": b64,
            },
        })

    prompt_text = GUIDELINE_PARSE_PROMPT
    if extra_text:
        prompt_text += f"\n\nAdditional text content extracted from spreadsheet:\n{extra_text}"

    content.append({"type": "text", "text": prompt_text})

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    response = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=8192,
        messages=[{"role": "user", "content": content}],
    )

    response_text = response.content[0].text.strip()

    # Clean up potential markdown wrapping
    if response_text.startswith("```"):
        response_text = response_text.split("\n", 1)[1]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        response_text = response_text.strip()

    parsed = json.loads(response_text)
    guideline = ParsedGuideline(**parsed)

    return guideline, all_images
